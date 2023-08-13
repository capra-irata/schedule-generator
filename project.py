from calendar import monthcalendar, setfirstweekday, SUNDAY
from os.path import expanduser
from string import ascii_uppercase
from sys import exit

from openpyxl import load_workbook, Workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Protection,
                             Side)


DESKTOP_PATH = expanduser('~/Desktop')
FILENAME_OUTPUT = 'Staff Schedule'
FILENAME_TEMPLATE = 'Staff Schedule - Template'

WEEKDAYS = ['Sunday',
            'Monday',
            'Tuesday',
            'Wednesday',
            'Thursday',
            'Friday',
            'Saturday'
            ]

BORDER = Side(border_style='thin')


def main():
    # Load template as worksheet to read user's desired schedule
    if template := load_template():
        print('Template found.\n\n'
              'A new schedule will be created for the given month and year...'
              )

    # If template doesn't exist, create it and then exit program
    else:
        print(f'Template file "{FILENAME_TEMPLATE}" not found.\n'
              'A blank template will be created on your desktop.\n\n'
              'Add staff names and shift hours to the template, '
              'then run this program again.\n'
              )

        input('Press ENTER to continue or CTRL+C to cancel...')

        template = create_template()
        template.save(f'{DESKTOP_PATH}\\{FILENAME_TEMPLATE}.xlsx')

        exit(1)

    # Store desired schedule as a dict of dicts in the form
    # {weekday: {hours: h, staff: s}}
    schedule = parse_template(template)

    # Prompt user for desired calendar month
    year, month = get_year_month()  # Unpack returned value

    print(f'\nThe file will be saved to your desktop as "{FILENAME_OUTPUT}".\n')

    input('Press ENTER to continue or CTRL+C to cancel...')

    # Write schedule to new workbook for the given calendar month
    wb = create_schedule(schedule, year, month)

    # Save to Excel file on user's desktop
    wb.save(f'{DESKTOP_PATH}\\{FILENAME_OUTPUT}.xlsx')

    exit(0)


def load_template():
    '''Load and return template worksheet if it exists'''
    try:
        wb = load_workbook(f'{DESKTOP_PATH}\\{FILENAME_TEMPLATE}.xlsx')
    except FileNotFoundError:
        return None
    else:
        return wb.active


def create_template():
    '''Create a blank template workbook and save file'''
    # Cell addresses and their corresponding text
    ROW_HEIGHT = 35
    COL_WIDTH = 14

    CATEGORIES = {'A2': 'O/N',
                  'A4': 'D/C',
                  'B2': 'Hours',
                  'B3': 'Staff',
                  'B4': 'Hours',
                  'B5': 'Staff'
                  }

    CATEGORY_COLOR = '00FFFF00'

    wb = Workbook()
    ws = wb.active

    # Enable sheet protection to lock certain cells from edits
    ws.protection.enable()

    # Add padding to worksheet for readability
    pad_cells(ws, row_height=ROW_HEIGHT, col_width=COL_WIDTH, end_row=5,
              end_col='I')

    # Write weekdays to top row of worksheet
    write_weekday_headers(ws, font_size=12, start_col=3)

    # Create borders for all relevant cells, minus headers
    for row in ws['A2':'I5']:
        for c in row:
            c.border = Border(left=BORDER,
                              right=BORDER,
                              top=BORDER,
                              bottom=BORDER
                              )

    # Two rows for O/N -> Staff and Hours
    ws.merge_cells('A2:A3')

    # Two rows for D/C -> Staff and Hours
    ws.merge_cells('A4:A5')

    for address, category in CATEGORIES.items():
        # Write categories to left two columns of worksheet
        c = ws[address]
        c.value = category

        # Style category cells
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.font = Font(name='Calibri', size=12, b=True)
        c.fill = PatternFill(fill_type='solid', start_color=CATEGORY_COLOR)

    # Adjust font for non-headers to be less prominent than headers
    for row in ws['C2':'I5']:
        for c in row:
            # Unlock non-header and non-category cells for user to edit
            c.protection = Protection(locked=False)

            c.alignment = Alignment(horizontal='left', vertical='center')
            c.font = Font(name='Calibri', size=12)

    return wb


def parse_template(ws):
    '''Read template and return organized data as a dict of dicts'''
    START_COL = 3
    END_COL = 9
    START_ROW = 2
    END_ROW = 5

    schedule = {}

    # Iterate over each column of template, from Sunday to Saturday,
    # ignoring table's headers
    for day, col in zip(WEEKDAYS, ws.iter_cols(
            min_col=START_COL, max_col=END_COL,
            min_row=START_ROW, max_row=END_ROW,
            values_only=True
        )):
        # Hours and staff strings are stored in separate cells in
        # template for UX, but will be combined for writing to schedule
        overnight = f'{col[0]} {col[1]}'
        coverage = f'{col[2]} {col[3]}'

        # Store each day's data with its corresponding weekday
        schedule[day] = {'O/N': overnight, 'D/C': coverage}

    return schedule


def get_year_month():
    '''Prompt user for desired year and month, validate, and return'''
    while True:
        m = input('Enter month (as a number): ').strip()
        try:
            m = int(m)
        except ValueError:
            print('Month must be a valid number.')
            continue
        else:
            if m < 1 or m > 12:
                print('Month must be between 1 and 12.')
                continue

        y = input('Enter year: ').strip()
        try:
            y = int(y)
        except ValueError:
            print('Year must be a valid number.')
            continue
        else:
            if y < 1:
                print('Year must be greater than 1.')
                continue

        # If no errors, exit loop
        break

    return y, m


def create_schedule(schedule, year, month):
    '''Write desired schedule out as a full calendar month and return
    workbook'''
    CELLS = 4  # Number of cells within each individual calendar date
    START_COL = 1
    START_ROW = 2  # Because row 1 contains headers

    DATE_COLOR = '00FFFF00'

    # Change calendar module to count weekdays starting from Sunday
    setfirstweekday(SUNDAY)  # SUNDAY is a constant from calendar module

    # Generate given month's calendar, formatted as a list of weeks,
    # each of which is a list of dates
    cal = monthcalendar(year, month)

    # Create empty workbook to write to
    wb = Workbook()
    ws = wb.active

    # Set up page layout for printing
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToHeight = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.print_area = 'A1:G25'

    # Add padding to worksheet for readability
    pad_cells(ws, row_height=25, col_width=20, end_row=25, end_col='G')

    # Write weekdays to top row of worksheet
    write_weekday_headers(ws, font_size=14)

    # Write schedule content
    for week_num, week in enumerate(cal):
        # Associate each date with its appropriate weekday and enumerate
        # so we can access the associated column
        for col, (date, weekday) in enumerate(
                zip(week, schedule), start=START_COL
            ):
            # Ignore "non-existent" dates from monthcalendar matrix
            if date == 0:
                continue

            # Each time we move to the next day of the week, reset the
            # row for the current week number
            row = (week_num * CELLS) + START_ROW

            # Write date
            c = ws.cell(row, col, value=f'{month:02d}/{date:02d}')

            # Style date cells
            c.alignment = Alignment(horizontal='right')
            c.border = Border(left=BORDER,
                              right=BORDER,
                              top=BORDER,
                              bottom=BORDER
                              )
            c.font = Font(name='Calibri', size=12, b=True)
            c.fill = PatternFill(fill_type='solid', start_color=DATE_COLOR)

            row += 1

            # shift is O/N or D/C, hours contains both hours and staff
            for shift, hours in schedule[weekday].items():
                c = ws.cell(row, col, value=f'{shift}: {hours}')

                # Style cells
                c.border = Border(left=BORDER,
                                  right=BORDER
                                  )
                c.font = Font(name='Calibri', size=10)

                row += 1

            # Final row is for listing extra staff not found on template
            c = ws.cell(row, col, value=f'T/C:')

            # Style cells
            c.border = Border(left=BORDER,
                              right=BORDER,
                              bottom=BORDER
                              )
            c.font = Font(name='Calibri', size=10)

    return wb


def pad_cells(ws, row_height, col_width, end_row, end_col:str):
    '''Add padding to cells of worksheet to improve readability'''
    # Set row heights
    for row in range(1, end_row + 1):
        ws.row_dimensions[row].height = row_height

    # Set column widths
    for col in ascii_uppercase:
        ws.column_dimensions[col].width = col_width

        # Stop iterating over alphabet when end reached
        if col == end_col:
            break


def write_weekday_headers(ws, font_size, start_col=1):
    '''Create and style headers that list days of week'''
    WEEKDAY_COLOR = '0EA1D2'

    # Write headers to first row of worksheet
    for i, weekday in enumerate(WEEKDAYS, start=start_col):
        c = ws.cell(row=1, column=i, value=weekday)

        # Style weekday cells
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = Border(left=BORDER,
                          right=BORDER,
                          top=BORDER,
                          bottom=BORDER
                          )
        c.font = Font(name='Calibri', size=font_size, b=True)
        c.fill = PatternFill(fill_type='solid', start_color=WEEKDAY_COLOR)


if __name__ == '__main__':
    main()